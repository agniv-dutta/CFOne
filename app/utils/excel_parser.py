"""Excel parsing utilities"""

import openpyxl
import xlrd
from typing import Dict, List, Any
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


def extract_data_from_excel(file_path: str) -> Dict[str, Any]:
    """
    Extract data from Excel spreadsheets

    Args:
        file_path: Path to Excel file

    Returns:
        Dictionary containing sheets data and metadata
    """
    result = {
        "sheets": [],
        "metadata": {
            "sheet_count": 0,
            "total_rows": 0
        }
    }

    try:
        # Try .xlsx format first (openpyxl)
        if file_path.endswith(".xlsx"):
            return _parse_xlsx(file_path)
        elif file_path.endswith(".xls"):
            return _parse_xls(file_path)
        else:
            # Try both
            try:
                return _parse_xlsx(file_path)
            except:
                return _parse_xls(file_path)

    except Exception as e:
        logger.error(f"Excel parsing failed: {str(e)}")

        if "password" in str(e).lower() or "encrypted" in str(e).lower():
            return {"error": "File is password-protected. Please provide an unencrypted version."}

        return {"error": f"Failed to parse Excel file: {str(e)}"}


def _parse_xlsx(file_path: str) -> Dict[str, Any]:
    """Parse .xlsx file using openpyxl"""
    result = {
        "sheets": [],
        "metadata": {
            "sheet_count": 0,
            "total_rows": 0
        }
    }

    workbook = openpyxl.load_workbook(file_path, data_only=True)
    result["metadata"]["sheet_count"] = len(workbook.sheetnames)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Get all rows
        rows = list(sheet.iter_rows(values_only=True))

        # Skip empty sheets
        if not rows or all(all(cell is None for cell in row) for row in rows):
            continue

        # Detect headers (assume first non-empty row)
        headers = None
        data_start_row = 0

        for idx, row in enumerate(rows):
            if any(cell is not None for cell in row):
                headers = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(row)]
                data_start_row = idx + 1
                break

        if headers is None:
            continue

        # Extract data rows
        data_rows = []
        for row in rows[data_start_row:]:
            # Skip empty rows
            if not any(cell is not None for cell in row):
                continue

            row_dict = {}
            for idx, cell in enumerate(row):
                if idx < len(headers):
                    # Convert datetime to string
                    if isinstance(cell, datetime):
                        cell = cell.strftime("%Y-%m-%d %H:%M:%S")
                    row_dict[headers[idx]] = cell

            if row_dict:  # Only add non-empty rows
                data_rows.append(row_dict)

        if data_rows:
            result["sheets"].append({
                "name": sheet_name,
                "headers": headers,
                "data": data_rows,
                "row_count": len(data_rows)
            })
            result["metadata"]["total_rows"] += len(data_rows)

    return result


def _parse_xls(file_path: str) -> Dict[str, Any]:
    """Parse .xls file using xlrd"""
    result = {
        "sheets": [],
        "metadata": {
            "sheet_count": 0,
            "total_rows": 0
        }
    }

    workbook = xlrd.open_workbook(file_path)
    result["metadata"]["sheet_count"] = workbook.nsheets

    for sheet in workbook.sheets():
        # Skip empty sheets
        if sheet.nrows == 0:
            continue

        # Get headers from first row
        headers = []
        for col in range(sheet.ncols):
            cell = sheet.cell(0, col)
            headers.append(str(cell.value) if cell.value else f"Column_{col}")

        # Extract data rows
        data_rows = []
        for row_idx in range(1, sheet.nrows):
            row_dict = {}
            has_data = False

            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                value = cell.value

                # Convert dates
                if cell.ctype == 3:  # Date type
                    try:
                        date_tuple = xlrd.xldate_as_tuple(value, workbook.datemode)
                        value = datetime(*date_tuple).strftime("%Y-%m-%d %H:%M:%S")
                    except:
                        pass

                row_dict[headers[col_idx]] = value
                if value:
                    has_data = True

            if has_data:
                data_rows.append(row_dict)

        if data_rows:
            result["sheets"].append({
                "name": sheet.name,
                "headers": headers,
                "data": data_rows,
                "row_count": len(data_rows)
            })
            result["metadata"]["total_rows"] += len(data_rows)

    return result
